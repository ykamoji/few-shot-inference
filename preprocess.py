import openpyxl as xl
import csv
import jinja2
import pathlib
import random
import os
import numpy as np
from tqdm import tqdm

# labels = [
#     {"label": 0, "answer": "Not Satisfactory"},
#     {"label": 1, "answer": "Satisfactory"},
# ]

questions = []
with open('Questions.csv', newline='') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for index, row in enumerate(reader):
        if index != 0:
            questions.extend(row)


wb = xl.load_workbook("health_articles_data.xlsx")

dataset = {}


def preprocess_inference():

    if os.path.exists('dataset.npy'):
        return np.load('dataset.npy').tolist()

    with pathlib.Path("template_inference.jinja2").open() as f:
        prompt_template = jinja2.Template(f.read())

    dataset = []
    
    for index in list(range(9)):
        # print("Active sheet: ", wb.active)

        # dataset[index] = []
        local_test_list = []

        wb._active_sheet_index = index
        sheet = wb.active

        dataDict = {}
        for row in tqdm(range(1, sheet.max_row)):
            data = list(sheet.iter_cols(1, sheet.max_column))
            example = data[21][row].value
            answer = data[18][row].value
            split = data[-1][row].value
            dataDict[row] = {"example": example, "question": questions[index], "answer": answer, "split": split}

            # print(f"Q:{questions[index]}\nEx:{example}\nAns:{answer}\nSplit:{split}\n\n")

        test_list = [index for index, data in dataDict.items() if data['split'] == 'test']
        # train_list = [index for index, data in dataDict.items() if data['split'] == 'train']

        for test_index in test_list:

            # train_index_1, train_index_2 = random.sample(train_list, 2) # train_list[:2]
            #
            # train_list.remove(train_index_1)
            # train_list.remove(train_index_2)

            # train_data_1, train_data_2 = dataDict[train_index_1], dataDict[train_index_2]

            test_data = dataDict[test_index]

            # print(f"{train_data_1}")
            # print(f"{train_data_2}")
            # print(f"{test_data}")

            # examples = [
            #     {"context": train_data_1["example"], "label": 1 if train_data_1["answer"] == "Satisfactory" else 0},
            #     {"context": train_data_2["example"], "label": 1 if train_data_2["answer"] == "Satisfactory" else 0},
            # ]

            prompt = prompt_template.render(
                context=test_data["example"],
                question=questions[index]
            )

            local_test_list.append({"input": prompt, "labels": 1 if test_data["answer"] == "Satisfactory" else 0})

        perc = 0.5
        local_test_list = random.sample(local_test_list, int(len(local_test_list) * perc))

        print(f"Completed {index + 1} criteria with {len(local_test_list)} testing dataset")

        dataset.extend(local_test_list)

    random.shuffle(dataset)
    print(f"Total {len(dataset)} testing dataset")

    np_dataset = np.array(dataset)

    np.save('dataset.npy', np_dataset)

    return dataset


def preprocess_training():

    with pathlib.Path("template_training.jinja2").open() as f:
        prompt_template = jinja2.Template(f.read())
    
    if os.path.exists('train.npy'):
        return np.load('train.npy', allow_pickle=True).tolist(), np.load('test.npy', allow_pickle=True).tolist()

    train_list, test_list = [], []

    for index in list(range(1)):
        # print("Active sheet: ", wb.active)
        wb._active_sheet_index = index
        sheet = wb.active

        local_train_list, local_test_list = [], []
        for row in tqdm(range(1, sheet.max_row)):
            data = list(sheet.iter_cols(1, sheet.max_column))
            context = data[21][row].value
            answer = data[18][row].value
            split = data[-1][row].value

            prompt = prompt_template.render(
                context=context,
                question=questions[index]
                # label=1 if answer == "Satisfactory" else 0
            )

            data = {"input": prompt, "label": 1 if answer == "Satisfactory" else 0}

            if split == 'train':
                local_train_list.append(data)
            else:
                local_test_list.append(data)

        perc = 1.0
        local_train_list = random.sample(local_train_list, int(len(local_train_list)*perc))
        local_test_list = random.sample(local_test_list, int(len(local_test_list)*perc))

        print(f"Completed {index+1} criteria with {len(local_train_list)} training dataset and "
              f"{len(local_test_list)} testing dataset")

        train_list.extend(local_train_list)
        test_list.extend(local_test_list)

    random.shuffle(train_list)
    random.shuffle(test_list)

    print(f"Total {len(train_list)} training dataset and " f"{len(test_list)} testing dataset")

    np_train_list = np.array(train_list)
    np_test_list = np.array(test_list)

    np.save('train.npy', np_train_list)
    np.save('test.npy', np_test_list)

    return train_list, test_list


def preprocess_few_shot_prompts():
    for index in list(range(5, 10)):

        wb._active_sheet_index = index
        sheet = wb.active

        dataDict = {}
        for row in tqdm(range(1, sheet.max_row)):
            data = list(sheet.iter_cols(1, sheet.max_column))
            example = data[21][row].value
            answer = data[18][row].value
            split = data[-1][row].value
            id = data[3][row].value
            title = data[11][row].value
            dataDict[row] = {"id": int(id), "title": title, "example": example,
                             "answer": 1 if answer == "Satisfactory" else 0, "split": split}

        test_list = [index for index, data in dataDict.items() if data['split'] == 'test']
        train_list = [index for index, data in dataDict.items() if data['split'] == 'train']

        train_positive, train_negative, test_positive, test_negative = 2, 2, 5, 5

        examples = []
        tests = []
        example_seq = 1
        while train_positive > 0 or train_negative > 0 or test_negative > 0 or test_positive > 0:

            train_indices = random.sample(train_list, 4)
            test_indices = random.sample(test_list, 10)

            for idx in train_indices + test_indices:
                if (idx in train_indices and
                    ((dataDict[idx]['answer'] == 1 and train_positive > 0) or (
                            dataDict[idx]['answer'] == 0 and train_negative > 0))) or \
                        (idx in test_indices and
                         (dataDict[idx]['answer'] == 1 and test_positive > 0) or (
                                 dataDict[idx]['answer'] == 0 and test_negative > 0)):

                    prompt_data = {
                        "title": dataDict[idx]['title'],
                        "news": dataDict[idx]['example'],
                        "output": dataDict[idx]['answer']
                    }

                    if idx in train_indices:

                        train_list.remove(idx)
                        prompt_data["idx"] = example_seq
                        example_seq += 1
                        examples.append(prompt_data)

                        if dataDict[idx]['answer'] == 1:
                            train_positive -= 1
                        else:
                            train_negative -= 1

                    else:

                        test_list.remove(idx)
                        prompt_data['id'] = dataDict[idx]['id']
                        tests.append(prompt_data)

                        if dataDict[idx]['answer'] == 1:
                            test_positive -= 1
                        else:
                            test_negative -= 1

        tests = sorted(tests, key=lambda data: data['output'])
        # examples = sorted(examples, key=lambda data: data['output'])

        for test in tests:

            modified_examples = []
            p_c, n_c = 1, 1
            for example in examples:
                if example['output'] == 0 and n_c == 1:
                    modified_examples.append(example)
                    n_c -= 1
                elif example['output'] == 1 and p_c == 1:
                    modified_examples.append(example)
                    p_c -= 1

            modified_examples[0]['idx'] = 1
            modified_examples[1]['idx'] = 2

            prompt = gen_prompt_2(questions[index], modified_examples, test)

            with open(f'results/criteria_{index + 1}_shot_2.txt', 'a') as file:
                file.write(str(test['id']) + "\n")
                file.write(str(test['output']) + "\n")
                file.write(prompt)
                file.write("\n\n")

            prompt = gen_prompt_4(questions[index], examples, test)

            with open(f'results/criteria_{index + 1}_shot_4.txt', 'a') as file:
                file.write(str(test['id']) + "\n")
                file.write(str(test['output']) + "\n")
                file.write(prompt)
                file.write("\n\n")


def gen_prompt_2(question, examples, test):
    def example_prompt(idx, title, news, output):
        return f'''
        Example {idx}:
        Title: {title}
        News Article: {news}
        Output: {output}
        '''

    prompt = f'''
    You are an expert evaluator tasked with assessing the quality of online health articles. You will evaluate each article based on a specific criterion and provide a concise evaluation. The criterion should be rated as either "Satisfactory" or "Non Satisfactory".

    Criterion for Evaluation:
    '{question}'

    {example_prompt(**examples[0])}
    
    {example_prompt(**examples[1])}
    
    Evaluate the following news article:
    Title: {test['title']}
    News Article: {test['news']}

    Output Format:
    Respond with two numbers, first either "1" for "Satisfactory" or "0" for "Not Satisfactory". Do not include any additional text.
    '''

    return prompt

def gen_prompt_4(question, examples, test):
    def example_prompt(idx, title, news, output):
        return f'''
        Example {idx}:
        Title: {title}
        News Article: {news}
        Output: {output}
        '''

    prompt = f'''
    You are an expert evaluator tasked with assessing the quality of online health articles. You will evaluate each article based on a specific criterion and provide a concise evaluation. The criterion should be rated as either "Satisfactory" or "Non Satisfactory".

    Criterion for Evaluation:
    '{question}'

    {example_prompt(**examples[0])}

    {example_prompt(**examples[1])}
    
    {example_prompt(**examples[2])}
    
    {example_prompt(**examples[3])}

    Evaluate the following news article:
    Title: {test['title']}
    News Article: {test['news']}

    Output Format:
    Respond with two numbers, first either "1" for "Satisfactory" or "0" for "Not Satisfactory". Do not include any additional text.
    '''

    return prompt


if __name__ == '__main__':
    preprocess_few_shot_prompts()

