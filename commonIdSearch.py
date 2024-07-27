from tqdm import tqdm
import openpyxl as xl

wb = xl.load_workbook("health_articles_data.xlsx")

idMap = {}
dataDict = {}
for index in list(range(10)):

    wb._active_sheet_index = index
    sheet = wb.active

    for row in tqdm(range(1, sheet.max_row)):
        data = list(sheet.iter_cols(1, sheet.max_column))
        id = int(data[3][row].value)
        example = data[21][row].value
        answer = data[18][row].value
        question = data[17][row].value
        title = data[11][row].value

        if id not in idMap.keys():
            if index == 0:
                idMap[id] = 1
                dataDict[id] = {"example": example, "questions": [question],
                                "title": title, "answers": [1 if answer == "Satisfactory" else 0]}
        else:
            idMap[id] += 1
            questions = dataDict[id]['questions'] + [question]
            answers = dataDict[id]['answers'] + [1 if answer == "Satisfactory" else 0]
            dataDict[id] = {"example": example, "questions": questions,
                            "title": title, "answers": answers}

print(idMap)

count = 0
for k,v in idMap.items():
    if v > 9:
        count += 1
        print(f"Id = {k}")
        print(f"Title:{dataDict[k]['title']}\nNews Article:{dataDict[k]['example']}")
        for i in range(len(dataDict[k]['questions'])):
            print(f"Criterion for Evaluation:\n{dataDict[k]['questions'][i]}")
            print(f"Output: {dataDict[k]['answers'][i]}")

        print("\n\n")

        if count > 5:
            break

